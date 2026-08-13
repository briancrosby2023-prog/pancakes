"""Build and persist the Operation Pancake evidence index."""

from __future__ import annotations

import json
import re
from dataclasses import replace
from pathlib import Path

from openpyxl import load_workbook

from operation_pancake.evidence.index import EvidenceIndex
from operation_pancake.evidence.models import (
    EvidenceLink,
    FieldProvenance,
    ReconciliationItem,
    SourceRecord,
)

WORKBOOK = Path("data/canonical/canonical_v1.9.xlsx")
INVENTORY = Path("data/research/progression_audit/progression_inventory.json")
HISTORICAL_CENTER = Path(
    "data/research/center_source_reconciliation/center_historical_reconciliation_pass1.json"
)


def _status(raw: str) -> tuple[str, str | None]:
    upper = raw.upper()
    if upper.startswith("COMPLETE"):
        return "COMPLETE", None
    if upper.startswith("PARTIAL"):
        return "PARTIAL", "Source registry identifies remaining evidence as unprocessed."
    if upper.startswith("REGISTERED"):
        return "UNPROCESSED", "Registered source has not been extracted."
    if upper.startswith("DUPLICATE"):
        return "NEEDS_REVIEW", "Duplicate/cross-check relationship requires reconciliation."
    if upper.startswith("EXTRACTED"):
        remaining = None if "COMPLETE" in upper else "Only the stated subset has been extracted."
        return ("COMPLETE" if remaining is None else "PARTIAL"), remaining
    return "NEEDS_REVIEW", "Legacy status requires review."


def _source_type(raw: str) -> str:
    upper = raw.upper()
    return next(
        (kind for kind in ("PDF", "IMAGE", "WORKBOOK", "CSV", "WEB", "API") if kind in upper),
        "OTHER",
    )


def _slug(value: str) -> str:
    return re.sub(r"[^A-Z0-9]+", "-", value.upper()).strip("-")


def build_evidence_index(root: Path) -> EvidenceIndex:
    """Build a deterministic index from authoritative repository evidence."""
    index = EvidenceIndex()
    workbook_path = root / WORKBOOK
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    rows = list(workbook["Source_Registry"].iter_rows(min_row=5, values_only=True))
    for row in rows:
        if not row[0]:
            continue
        source_id, filename, raw_type, domain, raw_status, url, notes, use = row[:8]
        extraction, remaining = _status(str(raw_status or ""))
        position = tuple(p for p in ("QB", "TE", "C") if re.search(rf"\b{p}\b", str(domain)))
        index.add_source(
            SourceRecord(
                source_id=str(source_id),
                source_name=str(domain),
                original_filename=str(filename),
                source_type=_source_type(str(raw_type)),
                category=str(domain),
                origin="LOCAL_REPOSITORY",
                positions=position,
                extraction_status=extraction,
                extraction_remaining=remaining,
                validation_status="WORKBOOK_REGISTERED",
                canonical_ingestion_status=str(use or "UNKNOWN"),
                research_use_status="USED" if "research" in str(use).lower() else "AVAILABLE",
                notes="; ".join(x for x in (str(notes or ""), str(url or "")) if x),
                provenance="REPOSITORY_CONFIRMED",
            )
        )

    # Preserve legacy IDs while recording the workbook-declared duplicate relationship.
    index.sources["SRC-C-RAW-002"] = replace(
        index.sources["SRC-C-RAW-002"],
        duplicate_of="SRC-C-RAW-001",
        notes=(
            f"{index.sources['SRC-C-RAW-002'].notes} Reconciled as a duplicate/cross-check "
            "of SRC-C-RAW-001; provenance retained."
        ),
    )

    # This known File Library item is intentionally distinct from the EA roster reference.
    index.add_source(
        SourceRecord(
            source_id="SRC-C-RAW-003",
            source_name="Partial raw Center CUT card source",
            original_filename="raw str centers2(2).pdf",
            source_type="PDF",
            category="CUT_CENTER_CARDS",
            origin="CHATGPT_FILE_LIBRARY",
            discovered_date="2026-08-13",
            positions=("C",),
            item_count=14,
            extraction_status="PARTIAL",
            extraction_remaining=(
                "Q-C-004-014: visually transcribe and validate Center CUT cards/rating panels "
                "on pages 4-14; do not infer rating vectors from weighted averages."
            ),
            validation_status="NEEDS_REVIEW",
            canonical_ingestion_status="NOT_INGESTED",
            research_use_status="HISTORICAL_CENTER_EVIDENCE",
            provenance="HISTORICALLY_RECOVERED",
            notes=(
                "Confirmed to exist in ChatGPT File Library. Historical ingestion was incomplete; "
                "this CUT source is not the EA base-roster reference dataset."
            ),
        )
    )
    index.add_source(
        SourceRecord(
            source_id="SRC-C-HIST-WB-001",
            source_name="Historical Madden 19 Center formula workbook",
            original_filename="Operation_Pancake_Madden19_Center_Formula.xlsx",
            source_type="WORKBOOK",
            category="HISTORICAL_CENTER_MODEL_RESEARCH",
            origin="CHATGPT_FILE_LIBRARY",
            discovered_date="2026-08-13",
            positions=("C",),
            player_card_coverage=(
                "Ashton Beers",
                "Justin Evans",
                "Bruce Mitchell",
                "Carson Hinzman",
                "Brady Small",
                "Coleton Price",
                "Landen Hatchett",
                "Lyndon Cooper",
                "Jake Guarnera",
                "Jake Renfro",
                "Levi Hubbard",
            ),
            item_count=12,
            extraction_status="COMPLETE",
            extraction_remaining=None,
            validation_status="HISTORICALLY_RECOVERED",
            canonical_ingestion_status="RESEARCH_ONLY",
            research_use_status="HISTORICAL_MODEL_METADATA",
            notes="Contains historical comparison and 53-player Madden Center model metadata.",
            provenance="HISTORICALLY_RECOVERED",
        )
    )

    historical = json.loads((root / HISTORICAL_CENTER).read_text(encoding="utf-8"))
    for number, observation in enumerate(historical["center_comparison"], start=1):
        record_id = f"HIST-C-{number:03d}"
        values = {
            **observation,
            "position": "C",
            "canonical": False,
            "evidence_class": "HISTORICAL_DERIVED_COMPARISON",
            "provenance_status": "HISTORICALLY_RECOVERED",
            "unextracted_cut_fields": "Complete CUT rating vector; validate from Q-C-004-014",
        }
        index.add_record("historical_center_observation", record_id, values)
        for source_id, relationship in (
            ("SRC-C-HIST-WB-001", "HISTORICAL_DERIVED_RESULT"),
            ("SRC-C-RAW-003", "UNDERLYING_CUT_CARD_SOURCE"),
        ):
            index.add_link(
                EvidenceLink(
                    link_id=f"LINK-{record_id}-{source_id}",
                    source_id=source_id,
                    target_type="historical_center_observation",
                    target_id=record_id,
                    relationship=relationship,
                    locator="Q-C-004-014" if source_id == "SRC-C-RAW-003" else None,
                )
            )
        for field in ("player", "weighted_average", "cut_ovr"):
            index.add_field_provenance(
                FieldProvenance(
                    provenance_id=f"PROV-{record_id}-{field}",
                    target_type="historical_center_observation",
                    target_id=record_id,
                    field_name=field,
                    value=values[field],
                    source_id="SRC-C-HIST-WB-001",
                    locator="Recovered historical Center comparison",
                    extraction_method="FILE_LIBRARY_RECONCILIATION",
                    validation_status="HISTORICALLY_RECOVERED",
                    confidence="historical_evidence",
                    provenance_status="HISTORICALLY_RECOVERED",
                    recorded_at="2026-08-13",
                )
            )

    index.add_record(
        "historical_model_result",
        "HIST-M19-CENTER-MODEL-001",
        {
            **historical["madden_center_model"],
            "canonical": False,
            "production_formula": False,
            "source_filename": "Operation_Pancake_Madden19_Center_Formula.xlsx",
        },
    )
    index.add_link(
        EvidenceLink(
            link_id="LINK-HIST-M19-CENTER-MODEL-001",
            source_id="SRC-C-HIST-WB-001",
            target_type="historical_model_result",
            target_id="HIST-M19-CENTER-MODEL-001",
            relationship="HISTORICAL_MODEL_SOURCE",
        )
    )
    index.add_link(
        EvidenceLink(
            link_id="LINK-HIST-M19-CENTER-MODEL-001-EA-ROSTER",
            source_id="SRC-RATE-001",
            target_type="historical_model_result",
            target_id="HIST-M19-CENTER-MODEL-001",
            relationship="EA_BASE_ROSTER_REFERENCE_POPULATION",
            notes="Reference evidence only; not a CUT card source.",
        )
    )
    index.add_record(
        "historical_model_result",
        "HIST-CFB-CENTER-CURVE-001",
        {
            **historical["cfb_fitted_curve"],
            "canonical": False,
            "production_formula": False,
            "population": 12,
        },
    )
    index.add_link(
        EvidenceLink(
            link_id="LINK-HIST-CFB-CENTER-CURVE-001",
            source_id="SRC-C-HIST-WB-001",
            target_type="historical_model_result",
            target_id="HIST-CFB-CENTER-CURVE-001",
            relationship="HISTORICAL_DERIVED_RESULT",
        )
    )

    inventory = json.loads((root / INVENTORY).read_text(encoding="utf-8"))
    cards_by_id = {card["card_id"]: card for card in inventory["canonical_cards"]}
    for card in inventory["canonical_cards"]:
        card_id = card["card_id"]
        source_id = card["source_id"]
        index.add_record("player_card", card_id, {**card, "canonical": True})
        index.add_link(
            EvidenceLink(
                link_id=f"LINK-CARD-{card_id}",
                source_id=source_id,
                target_type="player_card",
                target_id=card_id,
                relationship="CANONICAL_SOURCE",
                locator=card.get("source_locator"),
            )
        )
        fields = {"player": card["player"], "overall": card["overall"], **card["attributes"]}
        for field, value in fields.items():
            index.add_field_provenance(
                FieldProvenance(
                    provenance_id=f"PROV-{card_id}-{field}",
                    target_type="player_card",
                    target_id=card_id,
                    field_name=field,
                    value=value,
                    source_id=source_id,
                    locator=card.get("source_locator"),
                    extraction_method="CANONICAL_WORKBOOK_IMPORT",
                    validation_status=card["validation_status"],
                    confidence="validated",
                    provenance_status="DIRECTLY_OBSERVED",
                    version="canonical_v1.9",
                )
            )

    for number, candidate in enumerate(inventory["progression_candidates"], start=1):
        candidate_id = f"PROG-CAND-{number:04d}"
        index.add_record("progression_evidence", candidate_id, {**candidate, "canonical": False})
        for card_id in (candidate["lower_card_id"], candidate["upper_card_id"]):
            card = cards_by_id[card_id]
            index.add_link(
                EvidenceLink(
                    link_id=f"LINK-{candidate_id}-{card_id}",
                    source_id=card["source_id"],
                    target_type="progression_evidence",
                    target_id=candidate_id,
                    relationship="SUPPORTING_CARD_SOURCE",
                    locator=card.get("source_locator"),
                    notes=f"Evidence card {card_id}",
                )
            )

    research_root = root / "data/research"
    for path in sorted(research_root.rglob("*")):
        if path.suffix.lower() not in {".json", ".csv"}:
            continue
        relative = path.relative_to(root).as_posix()
        artifact_id = f"ART-{_slug(relative)}"
        text = path.read_text(encoding="utf-8")
        referenced_cards = sorted(set(re.findall(r"(?:QB|TE|C)-\d{4}", text)))
        index.add_record(
            "research_artifact",
            artifact_id,
            {
                "path": relative,
                "canonical": False,
                "research_only": True,
                "referenced_card_ids": referenced_cards,
            },
        )
        for source_id in sorted(set(re.findall(r"SRC-[A-Z0-9-]+", text)) & index.sources.keys()):
            index.add_link(
                EvidenceLink(
                    link_id=f"LINK-{artifact_id}-{source_id}",
                    source_id=source_id,
                    target_type="research_artifact",
                    target_id=artifact_id,
                    relationship="CITED_OR_EMBEDDED_SOURCE",
                )
            )

    for source in index.sources.values():
        if source.extraction_status in {"PARTIAL", "UNPROCESSED", "NEEDS_REVIEW"}:
            issue = (
                "PARTIALLY_EXTRACTED"
                if source.extraction_status == "PARTIAL"
                else "NEEDS_VALIDATION"
            )
            resolved_center_alias = source.source_id in {"SRC-C-RAW-001", "SRC-C-RAW-002"}
            index.add_queue_item(
                ReconciliationItem(
                    item_id=f"REC-{source.source_id}",
                    source_id=source.source_id,
                    affected_type="source",
                    affected_id=source.source_id,
                    issue_type=issue,
                    status="RESOLVED" if resolved_center_alias else "OPEN",
                    priority="HIGH" if source.source_id == "SRC-C-RAW-003" else "MEDIUM",
                    notes=source.extraction_remaining or "Source requires review.",
                    resolution=(
                        "Legacy source identity reconciled and retained; remaining visual CUT "
                        "extraction is tracked by SRC-C-RAW-003 / Q-C-004-014."
                        if resolved_center_alias
                        else None
                    ),
                    created_at="2026-08-13",
                    resolved_at="2026-08-13" if resolved_center_alias else None,
                )
            )
    return index


def write_evidence_artifacts(root: Path) -> EvidenceIndex:
    """Write deterministic index and audit JSON artifacts."""
    index = build_evidence_index(root)
    output = root / "data/evidence"
    output.mkdir(parents=True, exist_ok=True)
    (output / "source_index.json").write_text(
        json.dumps(index.as_dict(), indent=2, sort_keys=True) + "\n", encoding="utf-8"
    )
    (output / "ingestion_audit.json").write_text(
        json.dumps(index.audit(), indent=2, sort_keys=True) + "\n", encoding="utf-8"
    )
    return index
