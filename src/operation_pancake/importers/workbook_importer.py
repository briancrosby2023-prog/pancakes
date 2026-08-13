"""Canonical workbook importer for Operation Pancake."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


@dataclass(frozen=True, slots=True)
class WorkbookRecord:
    """One raw record imported from the canonical workbook."""

    sheet_name: str
    row_number: int
    values: dict[str, Any]

    @property
    def source_record(self) -> str:
        """Return stable workbook provenance for this record."""
        return f"{self.sheet_name}!{self.row_number}"


class WorkbookImporter:
    """Read Operation Pancake workbook data without modifying the source."""

    def __init__(self, workbook_path: str | Path) -> None:
        self.workbook_path = Path(workbook_path)

    def validate_source(self) -> None:
        """Validate that the workbook source exists and is supported."""
        if not self.workbook_path.exists():
            raise FileNotFoundError(
                f"Canonical workbook not found: {self.workbook_path}"
            )

        if not self.workbook_path.is_file():
            raise ValueError(
                f"Canonical workbook path is not a file: {self.workbook_path}"
            )

        if self.workbook_path.suffix.casefold() not in {".xlsx", ".xlsm"}:
            raise ValueError(
                "Canonical workbook must be an .xlsx or .xlsm file."
            )

    def sheet_names(self) -> list[str]:
        """Return worksheet names without changing the workbook."""
        self.validate_source()

        workbook = load_workbook(
            self.workbook_path,
            read_only=True,
            data_only=True,
        )

        try:
            return list(workbook.sheetnames)
        finally:
            workbook.close()

    @staticmethod
    def _headers(sheet) -> tuple[int, list[str | None]]:
        """Detect and normalize the worksheet's actual header row."""
        best_row_number = 1
        best_row = None
        best_score = -1

        for row_number, row in enumerate(
            sheet.iter_rows(
                min_row=1,
               max_row=25,
                values_only=True,
            ),
            start=1,
        ):
            normalized = [
                str(value).strip() if value is not None else None
                for value in row
            ]

            populated = [value for value in normalized if value]
            if not populated:
                continue

            score = len(populated)

            header_signals = {
                "Player",
                "OVR",
                "Archetype",
                "Program",
                "Position",
                "QB_ID",
                "Card_ID",
            }
            score += 10 * len(header_signals.intersection(populated))

            if score > best_score:
                best_score = score
                best_row_number = row_number
                best_row = row

        if best_row is None:
            return 1, []

        headers: list[str | None] = []

        for value in best_row:
            if value is None:
                headers.append(None)
                continue

            header = str(value).strip()
            headers.append(header if header else None)

        return best_row_number, headers

    def records(self, sheet_name: str) -> list[WorkbookRecord]:
        """Read non-empty rows from one worksheet with provenance."""
        self.validate_source()

        workbook = load_workbook(
            self.workbook_path,
            read_only=True,
            data_only=True,
        )

        try:
            if sheet_name not in workbook.sheetnames:
                raise KeyError(f"Worksheet not found: {sheet_name}")

            sheet = workbook[sheet_name]
            header_row, headers = self._headers(sheet)

            records: list[WorkbookRecord] = []

            for row_number, row in enumerate(
                sheet.iter_rows(
                    min_row=header_row + 1,
                    values_only=True,
                ),
                start=header_row + 1,
            ):
                values = {
                    header: value
                    for header, value in zip(headers, row, strict=False)
                    if header is not None
                }

                if not any(value is not None for value in values.values()):
                    continue

                records.append(
                    WorkbookRecord(
                        sheet_name=sheet_name,
                        row_number=row_number,
                        values=values,
                    )
                )

            return records
        finally:
            workbook.close()